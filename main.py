import pandas as pd
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import time

planilha = pd.read_excel('PRODUTOS 10-2023 - teste - Copia.xls', dtype=str)
#print(planilha)

driver = webdriver.Chrome()
driver.get("https://www.econeteditora.com.br//pis_cofins/index.php?form[a]=0")

user_input = driver.find_element(By.XPATH,"//input[@name='Log']")
user_input.send_keys('gdc47781')
time.sleep(3)
password_input = driver.find_element(By.XPATH,"//input[@name='Sen']")
password_input.send_keys('X32023')
time.sleep(10)

workbook = openpyxl.Workbook()
workbook.create_sheet("consulta")
sheet_consulta = workbook['consulta']

sheet_consulta['A1'].value = "Dt. Escrituação"
sheet_consulta['B1'].value = "Número"
sheet_consulta['C1'].value = "Mod."
sheet_consulta['D1'].value = "Nome Produto"
sheet_consulta['E1'].value = "Classificação"
sheet_consulta['F1'].value = "Valor Total Item"
sheet_consulta['G1'].value = 'Desconto'
sheet_consulta['H1'].value = 'Valor ajustado'
sheet_consulta['I1'].value = 'CONSULTA SIMPLES NACIONAL'

for linha in planilha.index:
    booking_date = (planilha.loc[linha, "Dt. Escrituação"]) 
    number = (planilha.loc[linha, "Número"])
    model = (planilha.loc[linha, "Mod."])
    product_name = (planilha.loc[linha, "Nome Produto"])
    ncm = (planilha.loc[linha, "Classificação"])
    total_value = (planilha.loc[linha, "Valor Total Item"])
    discount = (planilha.loc[linha, "Descontos"])
    ajusted_value = ''

    driver.switch_to.frame("ifram")
    ncm_input = driver.find_element(By.XPATH, "//input[@name='form[ncm]']")
    ncm_input.clear()
    ncm_input.send_keys(ncm)
    search_buttom = driver.find_element(By.XPATH, "//input[@value='Pesquisar']")
    search_buttom.click()
    time.sleep(1)
    select_ncm = driver.find_element(By.XPATH, "//input[@type='radio' and @name='form[ncm]']")
    time.sleep(1)
    select_ncm.click()
    time.sleep(1)

    consults = driver.find_elements(By.XPATH, "//*[@id='abas_internas']/div/div[1]/div[3]/table[1]/tbody/tr")
    time.sleep(1)

    #last_row = sheet_consulta.max_row + 1
    #column = 7

    for consult in consults:
        if consult.text.startswith('SIMPLES NACIONAL'):
            res_consult = consult.text
    
    sheet_consulta.append([booking_date, number, model, product_name, ncm, total_value, discount, ajusted_value, res_consult])
    '''
    for consult in consults:
        if consult.text.startswith('SIMPLES NACIONAL'):
            sheet_consulta.cell(row=last_row, column=column).value = consult.text
    '''
    '''
    for consult in consults:
        sheet_consulta.cell(row=last_row, column=column).value = consult.text
        #last_row += 1
        column += 1
        '''
    driver.switch_to.default_content()
    back_buttom = driver.find_element(By.XPATH, "//*[@id='link_voltar']")
    back_buttom.click()
    time.sleep(1)
    back_buttom.click()
    time.sleep(1)
    workbook.save('teste.xlsx')










#ncm_input = driver.find_element(driver.find_element(By.XPATH,"//input[@name='form[ncm]']"))
#time.sleep(2)



